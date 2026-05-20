import logging
from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import FraudThreat, ContextValidation, FraudAnalysisMetadata, FoodFraudExportResult

logger = logging.getLogger("vigia.food_fraud_exporter")

# ── Paleta Food Fraud ─────────────────────────────────────────────────────
C_HEADER_BG = "1F4E79"   # Azul oscuro
C_HEADER_FG = "FFFFFF"   # Blanco
C_UNCERTAIN_BG = "FFF3CD" # Ámbar claro para celdas inciertas
C_STAMP_FG = "808080"    # Gris para el sello

# Colores de Riesgo
C_RISK_EXTREMELY_HIGH = "FF0000" # Rojo
C_RISK_HIGH = "FF6600"           # Naranja
C_RISK_MEDIUM = "FFCC00"         # Amarillo
C_RISK_LOW = "00CC00"            # Verde

C_WHITE = "FFFFFF"
C_BLACK = "000000"
C_BORDER = "BDD7EE"

_thin = Border(
    left=Side(style="thin", color=C_BORDER),
    right=Side(style="thin", color=C_BORDER),
    top=Side(style="thin", color=C_BORDER),
    bottom=Side(style="thin", color=C_BORDER),
)
_wrap_left   = Alignment(wrap_text=True, vertical="top", horizontal="left")
_wrap_center = Alignment(wrap_text=True, vertical="top", horizontal="center")

def _style(cell, value, bg: str = C_WHITE, bold=False, center=False, color=C_BLACK, size=10):
    cell.value     = value
    cell.font      = Font(name="Calibri", bold=bold, size=size, color=color)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = _wrap_center if center else _wrap_left
    cell.border    = _thin

def _header(cell, text: str):
    cell.value     = text
    cell.font      = Font(name="Calibri", bold=True, size=11, color=C_HEADER_FG)
    cell.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
    cell.alignment = _wrap_center
    cell.border    = _thin

# ── Hoja 1: Análisis de Amenazas ──────────────────────────────────────────

_HEADERS = [
    "#",
    "Descripción de la Amenaza",
    "Efecto Potencial",
    "Causa(s) Potencial(es)",
    "Severidad",
    "Ocurrencia",
    "Nivel de Riesgo (NR)",
    "Nivel",
    "Controles Actuales",
    "Propuestas de Mitigación",
    "Decisión",
    "Responsable",
    "Fecha Límite",
    "Re-evaluación Sev.",
    "Re-evaluación Ocurr.",
    "NR Esperado"
]
_WIDTHS = [5, 40, 40, 40, 10, 12, 15, 20, 40, 40, 10, 20, 15, 15, 15, 15]

def _get_risk_colors(level: str):
    if level == "Extremely High":
        return C_RISK_EXTREMELY_HIGH, C_WHITE
    if level == "High":
        return C_RISK_HIGH, C_WHITE
    if level == "Medium":
        return C_RISK_MEDIUM, C_BLACK
    if level == "Low":
        return C_RISK_LOW, C_BLACK
    return C_WHITE, C_BLACK

def export_food_fraud_xlsx(
    threats: List[FraudThreat],
    company_name: str,
    process_area: str,
    evaluator_name: str,
    evaluation_date: str,
    gate_result: dict,
    analysis_metadata: dict,
    output_dir: Path
) -> FoodFraudExportResult:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Análisis de Amenazas"

    # Encabezado de la empresa
    ws.merge_cells("A1:P1")
    _style(ws["A1"], f"EVALUACIÓN DE VULNERABILIDAD AL FRAUDE ALIMENTARIO - {company_name} - {process_area}", bg=C_HEADER_BG, color=C_WHITE, bold=True, size=14, center=True)
    ws.row_dimensions[1].height = 40

    # Subencabezado
    ws.merge_cells("A2:P2")
    _style(ws["A2"], f"Evaluador: {evaluator_name} | Fecha: {evaluation_date} | Status: {gate_result.get('gate')} (Score: {gate_result.get('score')})", size=10, center=True)
    ws.row_dimensions[2].height = 20

    # Encabezados de columna
    for ci, h in enumerate(_HEADERS, 1):
        _header(ws.cell(row=4, column=ci), h)
    ws.row_dimensions[4].height = 35

    # Filas de datos
    for ri, t in enumerate(threats, start=5):
        bg = C_UNCERTAIN_BG if t.uncertain else C_WHITE
        risk_bg, risk_fg = _get_risk_colors(t.risk_level)

        values = [
            t.id,
            t.threat_description,
            t.potential_effect,
            "\n".join(t.potential_causes),
            t.severity,
            t.occurrence,
            t.risk_score,
            t.risk_level,
            t.current_controls,
            t.mitigation_proposals,
            "", # Decisión
            "", # Responsable
            "", # Fecha Límite
            "", # Re-eval Sev
            "", # Re-eval Ocurr
            ""  # NR Esperado
        ]

        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci)
            if ci == 7 or ci == 8: # Risk Score and Level
                _style(cell, val, bg=risk_bg, color=risk_fg, bold=True, center=True)
            elif ci == 5 or ci == 6: # Sev and Ocurr
                _style(cell, val, bg=bg, center=True)
            else:
                _style(cell, val, bg=bg)
        
        ws.row_dimensions[ri].height = 60

    # Anchos de columna
    for ci, w in enumerate(_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Sello de verificación
    last_row = 5 + len(threats) + 1
    ws.merge_cells(f"A{last_row}:P{last_row}")
    stamp_text = gate_result.get("xlsx_stamp", "Análisis generado por VIGÍA Food Fraud")
    _style(ws[f"A{last_row}"], stamp_text, color=C_STAMP_FG, size=10, bold=False)
    ws[f"A{last_row}"].font = Font(italic=True, color=C_STAMP_FG)

    # Hoja 2: Criterios de Evaluación
    ws2 = wb.create_sheet("Criterios de Evaluación")
    ws2["A1"] = "MATRIZ DE RIESGO DE FRAUDE (GFSI/FSMA)"
    ws2["A1"].font = Font(bold=True, size=12)
    
    # Simple representation of the matrix for reference
    matrix_headers = ["", "A (Frequent)", "B (Likely)", "C (Occasional)", "D (Seldom)", "E (Unlikely)"]
    for ci, h in enumerate(matrix_headers, 1):
        _header(ws2.cell(row=3, column=ci), h)
    
    matrix_rows = [
        ("I (Catastrophic)", 1, 2, 6, 8, 12),
        ("II (Critical)", 3, 4, 7, 11, 15),
        ("III (Moderate)", 5, 9, 10, 14, 16),
        ("IV (Negligible)", 13, 17, 18, 19, 20)
    ]
    
    for ri, m_row in enumerate(matrix_rows, 4):
        for ci, val in enumerate(m_row, 1):
            cell = ws2.cell(row=ri, column=ci)
            if ci == 1:
                _header(cell, val)
            else:
                score = int(val)
                level = "Extremely High" if score <= 4 else "High" if score <= 8 else "Medium" if score <= 14 else "Low"
                rbg, rfg = _get_risk_colors(level)
                _style(cell, val, bg=rbg, color=rfg, center=True, bold=True)

    # Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_company = "".join(c if c.isalnum() else "_" for c in company_name)[:20]
    safe_area = "".join(c if c.isalnum() else "_" for c in process_area)[:20]
    file_name = f"food_fraud_{safe_company}_{safe_area}_{timestamp}.xlsx"
    file_path = output_dir / file_name
    wb.save(file_path)

    high_priority_rows = sum(1 for t in threats if t.risk_score <= 8)

    return FoodFraudExportResult(
        file_path=str(file_path),
        file_name=file_name,
        download_url=f"/outputs/{file_name}",
        total_rows=len(threats),
        high_priority_rows=high_priority_rows,
        stamp_applied=stamp_text
    )
