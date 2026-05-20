import logging
from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import DefenseThreat, FoodDefenseExportResult

logger = logging.getLogger("vigia.food_defense_exporter")

# ── Paleta Food Defense (MRO) ─────────────────────────────────────────────
C_HEADER_BG = "D9E1F2"   # Azul claro grisáceo
C_TITLE_BG = "1F4E79"    # Azul oscuro
C_WHITE = "FFFFFF"
C_BLACK = "000000"
C_BORDER = "000000"

# Colores de Riesgo
C_RISK_HIGH = "FF0000"   # Rojo
C_RISK_MEDIUM = "FFFF00" # Amarillo
C_RISK_LOW = "00B050"    # Verde

_thin = Border(
    left=Side(style="thin", color=C_BORDER),
    right=Side(style="thin", color=C_BORDER),
    top=Side(style="thin", color=C_BORDER),
    bottom=Side(style="thin", color=C_BORDER),
)

_medium_top_bottom = Border(
    left=Side(style="thin", color=C_BORDER),
    right=Side(style="thin", color=C_BORDER),
    top=Side(style="medium", color=C_BORDER),
    bottom=Side(style="medium", color=C_BORDER),
)

_wrap_center = Alignment(wrap_text=True, vertical="center", horizontal="center")
_wrap_top_left = Alignment(wrap_text=True, vertical="top", horizontal="left")

def _style(cell, value, bg: str = C_WHITE, bold=False, center=True, color=C_BLACK, size=10, border=_thin):
    cell.value     = value
    cell.font      = Font(name="Arial", bold=bold, size=size, color=color)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = _wrap_center if center else _wrap_top_left
    cell.border    = border

def _get_risk_bg(level: str):
    if level == "High": return C_RISK_HIGH
    if level == "Medium": return C_RISK_MEDIUM
    return C_RISK_LOW

def export_food_defense_xlsx(
    threats: List[DefenseThreat],
    company_name: str,
    process_area: str,
    output_dir: Path
) -> FoodDefenseExportResult:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formato MRO"

    # 1. Título principal
    ws.merge_cells("A2:M2")
    _style(ws["A2"], "Análisis de Amenazas contra actos dolosos, sabotaje, contaminación intencional bajo MRO", bg=C_WHITE, bold=True, size=12)
    
    # 2. Encabezado de proceso y fecha
    ws.merge_cells("A3:G3")
    _style(ws["A3"], f"Proceso: {process_area} ({company_name})", center=False, bold=True)
    ws.merge_cells("H3:M3")
    _style(ws["H3"], f"Fecha: {datetime.now().strftime('%d/%m/%Y')}", center=False, bold=True)
    
    ws["A4"] = "Utilice diagrama de flujo"
    ws["A4"].font = Font(name="Arial", italic=True, size=10)

    # 3. Encabezados de tabla (Fila 5 y 6)
    headers_top = [
        ("A5:A6", "Descripción de la Amenaza. Lo que hace que el riesgo este presente. Considere operaciones sensibles, en donde los materiales y productos se encuentren expuestos."),
        ("B5:B6", "Efecto Potencial. Que impacto puede tener en productos, procesos, marcas de la compañía, etc."),
        ("C5:C6", "Causa(s) Potencial(es) \nde la Falla. Pueden existir varias causas. Describa cada una."),
        ("D5:F5", "Evaluacion del \nRiesgo"),
        ("G5:G6", "¿Se cuenta con algún control actualmente? Selección de personal, control de accesos, vigilancia de área, CCTV, otros, describa:"),
        ("H5:H6", "Propuestas de mitigación. Mayor supervisión, limitación de acceso, evaluación de confianza del personal, supervisión cercana, CCTV dirigido, etc."),
        ("I5:I6", "Decisiones para las propuestas: (A) Aprobada, (AM) Aprobada con modificaciones, (R) Rechazada"),
        ("J5:J6", "Responsables y fechas"),
        ("K5:M5", "Re-evaluacion del Riesgo (Nivel de riesgo residual esperado)")
    ]
    
    for r, text in headers_top:
        if ":" in r:
            ws.merge_cells(r)
            cell = ws[r.split(":")[0]]
        else:
            cell = ws[r]
        _style(cell, text, bg=C_HEADER_BG, bold=True, size=9)

    # Sub-encabezados (Fila 6)
    _style(ws["D6"], "Sev.", bg=C_HEADER_BG, bold=True, size=9)
    _style(ws["E6"], "Ocurr.", bg=C_HEADER_BG, bold=True, size=9)
    _style(ws["F6"], "Nivel Riesgo", bg=C_HEADER_BG, bold=True, size=9)
    _style(ws["K6"], "Sev.", bg=C_HEADER_BG, bold=True, size=9)
    _style(ws["L6"], "Ocurr.", bg=C_HEADER_BG, bold=True, size=9)
    _style(ws["M6"], "Nivel Riesgo esperado", bg=C_HEADER_BG, bold=True, size=9)

    # 4. Filas de datos
    for ri, t in enumerate(threats, start=7):
        # Merge de las primeras 3 columnas para que se vea como el original
        # El original no hace merge horizontal entre columnas, sino vertical si hay varios items.
        # Aquí seguiremos el flujo de una amenaza por fila.
        
        _style(ws.cell(row=ri, column=1), t.threat_description, center=False)
        _style(ws.cell(row=ri, column=2), t.potential_effect, center=False)
        _style(ws.cell(row=ri, column=3), "\n".join(t.potential_causes), center=False)
        
        # Evaluación inicial
        _style(ws.cell(row=ri, column=4), t.severity)
        _style(ws.cell(row=ri, column=5), t.occurrence)
        _style(ws.cell(row=ri, column=6), t.risk_score, bg=_get_risk_bg(t.risk_level), bold=True)
        
        _style(ws.cell(row=ri, column=7), t.current_controls, center=False)
        _style(ws.cell(row=ri, column=8), t.mitigation_proposals, center=False)
        _style(ws.cell(row=ri, column=9), "") # Decisiones
        _style(ws.cell(row=ri, column=10), "") # Responsables
        
        # Evaluación residual
        _style(ws.cell(row=ri, column=11), t.residual_severity)
        _style(ws.cell(row=ri, column=12), t.residual_occurrence)
        _style(ws.cell(row=ri, column=13), t.residual_risk_score, bg=_get_risk_bg(t.residual_risk_level), bold=True)
        
        ws.row_dimensions[ri].height = 80

    # 5. Nota al pie
    last_row = 7 + len(threats) + 1
    ws.merge_cells(f"A{last_row}:M{last_row}")
    note = "NR(Nivel de riesgo): Resulta de multiplicar los índices de Severidad x Ocurrencia = NR. Su valor puede ir desde 1 hasta 12. Las operaciones que hayan obtenido los mayores valores de NR serán los primeros en aplicar Acciones recomendadas"
    _style(ws[f"A{last_row}"], note, center=False, size=9)

    # Anchos de columna
    widths = [40, 30, 30, 5, 7, 12, 35, 40, 15, 20, 5, 7, 12]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Hoja de criterios (resumida)
    ws2 = wb.create_sheet("Criterios de evaluación")
    ws2["A1"] = "MANEJO DE RIESGO OPERACIONAL: CRITERIOS DE EVALUACION DE RIESGOS"
    ws2["A1"].font = Font(bold=True, size=12)
    
    # Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_company = "".join(c if c.isalnum() else "_" for c in company_name)[:20]
    file_name = f"food_defense_{safe_company}_{timestamp}.xlsx"
    file_path = output_dir / file_name
    wb.save(file_path)

    high_priority_rows = sum(1 for t in threats if t.risk_score >= 9)

    return FoodDefenseExportResult(
        file_path=str(file_path),
        file_name=file_name,
        download_url=f"/outputs/{file_name}",
        total_rows=len(threats),
        high_priority_rows=high_priority_rows
    )
