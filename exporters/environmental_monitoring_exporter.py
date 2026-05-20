import logging
from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import EnvironmentalThreat, EnvironmentalExportResult

logger = logging.getLogger("vigia.environmental_monitoring_exporter")

# ── Paleta ──────────────────────────────────────────────────────────────────
C_HEADER_BG = "C0C0C0"
C_TITLE_BG = "FFFFFF"
C_WHITE = "FFFFFF"
C_BLACK = "000000"
C_BORDER = "000000"

C_RISK_HIGH = "FF0000"
C_RISK_MEDIUM = "FFFF00"
C_RISK_LOW = "00B050"

_thin = Border(
    left=Side(style="thin", color=C_BORDER),
    right=Side(style="thin", color=C_BORDER),
    top=Side(style="thin", color=C_BORDER),
    bottom=Side(style="thin", color=C_BORDER),
)

_wrap_center = Alignment(wrap_text=True, vertical="center", horizontal="center")
_wrap_top_left = Alignment(wrap_text=True, vertical="top", horizontal="left")


def _style(cell, value, bg=C_WHITE, bold=False, center=True, color=C_BLACK, size=10):
    cell.value = value
    cell.font = Font(name="Arial", bold=bold, size=size, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = _wrap_center if center else _wrap_top_left
    cell.border = _thin


def _get_risk_bg(level: str):
    if level == "High": return C_RISK_HIGH
    if level == "Medium": return C_RISK_MEDIUM
    return C_RISK_LOW


def export_environmental_monitoring_xlsx(
    threats: List[EnvironmentalThreat],
    company_name: str,
    process_area: str,
    gate_result: dict,
    output_dir: Path
) -> EnvironmentalExportResult:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Identificacion de servicios"

    # 1. Titulo principal
    ws.merge_cells("A1:N1")
    _style(ws["A1"], "Analisis de Riesgo Monitoreo Ambiental Microbiologico", bg=C_TITLE_BG, bold=True, size=18, center=True)
    ws.row_dimensions[1].height = 60

    # 2. Encabezado proceso y fecha
    ws.merge_cells("A2:E2")
    _style(ws["A2"], f"Proceso: {process_area} ({company_name})", center=False, bold=True, size=11)
    ws.merge_cells("I2:N2")
    _style(ws["I2"], f"Fecha: {datetime.now().strftime('%d/%m/%Y')}", center=False, bold=True, size=11)

    # 3. Encabezados de tabla (Fila 4 y 5)
    _style(ws.cell(row=4, column=1), "Area, Proceso o Servicio", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=2), "CATEGORIZACION DE AREA", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=3), "Riesgos microbiologicos: Patogenos significativos, Coliformes, Cuenta total, etc. (Por area)", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=4), "Que activa el riesgo: Aire ambiental, personal que tiene contacto con el producto, aire asociado al proceso, limpieza con agua no controlada, acumulacion de agua en canaletas, registros de cañeria, condensacion, etc. Describa:", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=5), "Habilidades. Describa como se limpia el equipo o area, acciones de saneamiento, actividades de mantenimiento, filtracion de aire, monitoreo, etc.", bg=C_HEADER_BG, bold=True, size=9, center=True)
    ws.merge_cells("F4:H4")
    _style(ws["F4"], "Evaluacion del Riesgo", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=9), "Requiere medidas especiales de control? Mayor rigurosidad de ejecucion, supervision, cerrar frecuencias, mayor verificacion, inspecciones, etc. Describa", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=10), "Establezca frecuencia de verificacion en funcion del nivel de riesgo determinado", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=11), "Responsables y fechas", bg=C_HEADER_BG, bold=True, size=9, center=True)
    ws.merge_cells("L4:N4")
    _style(ws["L4"], "Re-evaluacion del Riesgo (Nivel de riesgo residual esperado)", bg=C_HEADER_BG, bold=True, size=9, center=True)

    # Sub-encabezados (Fila 5)
    _style(ws.cell(row=5, column=6), "Sev.", bg=C_HEADER_BG, bold=True, size=9)
    _style(ws.cell(row=5, column=7), "Ocurr.", bg=C_HEADER_BG, bold=True, size=9)
    _style(ws.cell(row=5, column=8), "Nivel Riesgo", bg=C_HEADER_BG, bold=True, size=9)
    _style(ws.cell(row=5, column=12), "Sev.", bg=C_HEADER_BG, bold=True, size=9)
    _style(ws.cell(row=5, column=13), "Ocurr.", bg=C_HEADER_BG, bold=True, size=9)
    _style(ws.cell(row=5, column=14), "Nivel Riesgo esperado", bg=C_HEADER_BG, bold=True, size=9)

    # 4. Filas de datos
    for ri, t in enumerate(threats, start=6):
        _style(ws.cell(row=ri, column=1), t.area, center=False)
        _style(ws.cell(row=ri, column=2), t.area_category, center=True)
        _style(ws.cell(row=ri, column=3), t.microbiological_risk, center=False)
        _style(ws.cell(row=ri, column=4), t.risk_activators, center=False)
        _style(ws.cell(row=ri, column=5), t.sanitation_abilities, center=False)

        _style(ws.cell(row=ri, column=6), t.severity)
        _style(ws.cell(row=ri, column=7), t.occurrence)
        _style(ws.cell(row=ri, column=8), t.risk_score, bg=_get_risk_bg(t.risk_level), bold=True)

        _style(ws.cell(row=ri, column=9), t.special_controls, center=False)
        _style(ws.cell(row=ri, column=10), t.verification_frequency)
        _style(ws.cell(row=ri, column=11), "")

        _style(ws.cell(row=ri, column=12), t.residual_severity)
        _style(ws.cell(row=ri, column=13), t.residual_occurrence)
        _style(ws.cell(row=ri, column=14), t.residual_risk_score, bg=_get_risk_bg(t.residual_risk_level), bold=True)

        ws.row_dimensions[ri].height = 80

    # 5. Nota al pie
    last_row = 6 + len(threats) + 1
    ws.merge_cells(f"A{last_row}:H{last_row}")
    note = "NR(Nivel de riesgo): Resulta de multiplicar los indices de Severidad x Ocurrencia = NR. Su valor puede ir desde 1 hasta 12. Las operaciones que hayan obtenido los mayores valores de NR seran los primeros en aplicar Acciones recomendadas"
    _style(ws[f"A{last_row}"], note, center=False, size=9)

    ws.merge_cells(f"I{last_row}:N{last_row}")
    note2 = "Nota: Si el resultante sobre el nivel de riesgo es medio alto, o alto, se debera evaluar el incremento de la rigurosidad del metodo operativo, o la supervision o la verificacion."
    _style(ws[f"I{last_row}"], note2, center=False, size=9)

    # Anchos de columna
    widths = [30, 22, 30, 35, 30, 6, 6, 10, 35, 20, 15, 6, 6, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Hoja de criterios
    ws2 = wb.create_sheet("Criterios de evaluacion")
    ws2.merge_cells("A1:G1")
    _style(ws2["A1"], "MANEJO DE RIESGO OPERACIONAL: CRITERIOS DE EVALUACION DE RIESGOS", bg=C_TITLE_BG, bold=True, size=12)
    ws2.row_dimensions[1].height = 30

    # Matriz de riesgo simple
    matrix_headers = ["", "Raro (1)", "Posible (2)", "Frecuente (3)"]
    for ci, h in enumerate(matrix_headers, 1):
        _style(ws2.cell(row=3, column=ci), h, bg=C_HEADER_BG, bold=True)

    matrix_rows = [
        ("Critico (4)", 4, 8, 12),
        ("Significativo (3)", 3, 6, 9),
        ("Menor (2)", 2, 4, 6),
        ("Insignificante (1)", 1, 2, 3),
    ]

    for ri, m_row in enumerate(matrix_rows, 4):
        for ci, val in enumerate(m_row, 1):
            cell = ws2.cell(row=ri, column=ci)
            if ci == 1:
                _style(cell, val, bg=C_HEADER_BG, bold=True)
            else:
                score = int(val)
                level = "High" if score >= 9 else "Medium" if score >= 4 else "Low"
                _style(cell, val, bg=_get_risk_bg(level), bold=True)

    # Sello de verificacion
    stamp_row = 6 + len(threats) + 2
    ws.merge_cells(f"A{stamp_row}:N{stamp_row}")
    stamp_text = gate_result.get("xlsx_stamp", "Analisis generado por VIGIA Monitoreo Ambiental")
    _style(ws[f"A{stamp_row}"], stamp_text, color="808080", size=10, center=True)

    # Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_company = "".join(c if c.isalnum() else "_" for c in company_name)[:20]
    safe_area = "".join(c if c.isalnum() else "_" for c in process_area)[:20]
    file_name = f"monitoreo_ambiental_{safe_company}_{safe_area}_{timestamp}.xlsx"
    file_path = output_dir / file_name
    wb.save(file_path)

    high_priority_rows = sum(1 for t in threats if t.risk_score >= 9)

    return EnvironmentalExportResult(
        file_path=str(file_path),
        file_name=file_name,
        download_url=f"/outputs/{file_name}",
        total_rows=len(threats),
        high_priority_rows=high_priority_rows
    )
