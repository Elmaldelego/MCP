import logging
from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import AllergenThreat, AllergenExportResult

logger = logging.getLogger("vigia.allergen_management_exporter")

C_HEADER_BG = "95B3D7"
C_TITLE_BG = "FFFFFF"
C_WHITE = "FFFFFF"
C_BLACK = "000000"
C_BORDER = "000000"

C_RISK_EXTREME = "FF0000"
C_RISK_HIGH = "FF6600"
C_RISK_MEDIUM = "FFCC00"
C_RISK_LOW = "00CC00"

_thin = Border(
    left=Side(style="thin", color=C_BORDER),
    right=Side(style="thin", color=C_BORDER),
    top=Side(style="thin", color=C_BORDER),
    bottom=Side(style="thin", color=C_BORDER),
)
_wrap_center = Alignment(wrap_text=True, vertical="center", horizontal="center")
_wrap_top_left = Alignment(wrap_text=True, vertical="top", horizontal="left")


def _style(cell, value, bg=C_WHITE, bold=False, center=True, color=C_BLACK, size=10):
    cell.value = value
    cell.font = Font(name="Arial", bold=bold, size=size, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = _wrap_center if center else _wrap_top_left
    cell.border = _thin


def _get_risk_bg(level: str):
    if level == "Extremely High": return C_RISK_EXTREME
    if level == "High": return C_RISK_HIGH
    if level == "Medium": return C_RISK_MEDIUM
    return C_RISK_LOW


def _get_risk_fg(level: str):
    if level in ("Extremely High", "High"): return C_WHITE
    return C_BLACK


def export_allergen_management_xlsx(
    threats: List[AllergenThreat],
    company_name: str,
    process_area: str,
    evaluator_name: str,
    evaluation_date: str,
    materials_involved: str,
    gate_result: dict,
    output_dir: Path
) -> AllergenExportResult:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Identificacion de servicios"

    # 1. Titulo
    ws.merge_cells("A1:J1")
    _style(ws["A1"], f"GRUPO DELCEN Analisis de Riesgo Gestion de alergenos - {company_name}", bg=C_TITLE_BG, bold=True, size=14, center=True)
    ws.row_dimensions[1].height = 30

    # 2. Proceso y fecha
    ws.merge_cells("A2:D2")
    _style(ws["A2"], f"Proceso: {process_area}", center=False, bold=True, size=11)
    ws.merge_cells("H2:J2")
    _style(ws["H2"], f"Fecha: {evaluation_date}", center=False, bold=True, size=11)

    # 3. Participantes y materiales
    ws.merge_cells("A3:D3")
    _style(ws["A3"], f"Participantes del equipo evaluador: {evaluator_name}", center=False, size=10)
    ws.merge_cells("H3:J3")
    _style(ws["H3"], f"Materiales y Productos involucrados: {materials_involved}", center=False, size=10)

    # 4. Encabezados (Fila 4 y 5)
    _style(ws.cell(row=4, column=1), "Area o Proceso", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=2), "Alergenos involucrados: Describa el alergeno o la combinacion de los mismos. Describa la forma de los alergenos.", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=3), "Descripcion de riesgos: Caracterizacion del riesgo. Como se puede presentar? Describa el impacto.", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=4), "Controles Actuales: Ej. Metodos operacionales, programas de prerrequisitos, control sobre material de etiquetado, etc.", bg=C_HEADER_BG, bold=True, size=9, center=True)
    ws.merge_cells("E4:G4")
    _style(ws["E4"], "Evaluacion del Riesgo", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=8), "Requiere medidas especiales de control? Mayor rigurosidad de ejecucion, supervision, cerrar frecuencias, mayor verificacion, inspecciones, etc. Describa", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=9), "Establezca metodo y frecuencia de verificacion", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=10), "Responsables y fechas", bg=C_HEADER_BG, bold=True, size=9, center=True)

    _style(ws.cell(row=5, column=5), "Sev.", bg=C_HEADER_BG, bold=True, size=9)
    _style(ws.cell(row=5, column=6), "Ocurr.", bg=C_HEADER_BG, bold=True, size=9)
    _style(ws.cell(row=5, column=7), "Nivel Riesgo", bg=C_HEADER_BG, bold=True, size=9)

    # 5. Filas de datos
    for ri, t in enumerate(threats, start=6):
        _style(ws.cell(row=ri, column=1), t.area, center=False)
        _style(ws.cell(row=ri, column=2), t.allergens_involved, center=False)
        _style(ws.cell(row=ri, column=3), t.risk_description, center=False)
        _style(ws.cell(row=ri, column=4), t.current_controls, center=False)

        _style(ws.cell(row=ri, column=5), t.severity)
        _style(ws.cell(row=ri, column=6), t.occurrence)
        risk_bg = _get_risk_bg(t.risk_level)
        risk_fg = _get_risk_fg(t.risk_level)
        _style(ws.cell(row=ri, column=7), t.risk_score, bg=risk_bg, bold=True, color=risk_fg)

        _style(ws.cell(row=ri, column=8), t.special_controls, center=False)
        _style(ws.cell(row=ri, column=9), t.verification_frequency, center=False)
        _style(ws.cell(row=ri, column=10), "")

        ws.row_dimensions[ri].height = 100

    # 6. Notas al pie
    last_data_row = 6 + len(threats)
    nr_row = last_data_row + 1
    ws.merge_cells(f"A{nr_row}:D{nr_row}")
    note1 = "NR (Nivel de riesgo): Resulta de multiplicar los indices de Severidad x Ocurrencia = NR. Las operaciones que hayan obtenido los mayores valores de NR seran los primeros en aplicar acciones recomendadas."
    _style(ws[f"A{nr_row}"], note1, center=False, size=9)

    ws.merge_cells(f"E{nr_row}:G{nr_row}")
    note2 = "Los criterios de evaluacion debe ser aprobada por el equipo de seguridad de los alimentos."
    _style(ws[f"E{nr_row}"], note2, center=False, size=9)

    ws.merge_cells(f"H{nr_row}:J{nr_row}")
    note3 = "Confirme que su matriz de Severidad y Ocurrencia define claramente los criterios."
    _style(ws[f"H{nr_row}"], note3, center=False, size=9)

    # Anchos de columna
    widths = [26, 24, 35, 32, 6, 7, 10, 32, 22, 18]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Hoja de criterios
    ws2 = wb.create_sheet("Criterios de evaluacion")
    ws2.merge_cells("B1:G1")
    _style(ws2["B1"], "MANEJO DE RIESGO OPERACIONAL: CRITERIOS DE EVALUACION DE RIESGOS", bg=C_TITLE_BG, bold=True, size=12)

    # Sello
    stamp_row = nr_row + 1
    ws.merge_cells(f"A{stamp_row}:J{stamp_row}")
    stamp_text = gate_result.get("xlsx_stamp", "Analisis generado por VIGIA Gestion de Alergenos")
    _style(ws[f"A{stamp_row}"], stamp_text, color="808080", size=10, center=True)

    # Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_company = "".join(c if c.isalnum() else "_" for c in company_name)[:20]
    safe_area = "".join(c if c.isalnum() else "_" for c in process_area)[:20]
    file_name = f"alergenos_{safe_company}_{safe_area}_{timestamp}.xlsx"
    file_path = output_dir / file_name
    wb.save(file_path)

    high_priority_rows = sum(1 for t in threats if t.risk_score <= 8)

    return AllergenExportResult(
        file_path=str(file_path),
        file_name=file_name,
        download_url=f"/outputs/{file_name}",
        total_rows=len(threats),
        high_priority_rows=high_priority_rows
    )