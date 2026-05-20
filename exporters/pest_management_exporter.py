import logging
from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import PestThreat, PestExportResult

logger = logging.getLogger("vigia.pest_management_exporter")

# ── Paleta ──────────────────────────────────────────────────────────────────
C_HEADER_BG = "C0C0C0"
C_TITLE_BG = "FFFFFF"
C_WHITE = "FFFFFF"
C_BLACK = "000000"
C_BORDER = "000000"

C_RISK_EXTREME = "000000"
C_RISK_CRITICAL = "FF0000"
C_RISK_HIGH = "FF6600"
C_RISK_MODERATE = "FFFF00"
C_RISK_LOW = "92D050"
C_RISK_MINIMAL = "00B050"

_thin = Border(
    left=Side(style="thin", color=C_BORDER),
    right=Side(style="thin", color=C_BORDER),
    top=Side(style="thin", color=C_BORDER),
    bottom=Side(style="thin", color=C_BORDER),
)

_wrap_center = Alignment(wrap_text=True, vertical="center", horizontal="center")
_wrap_top_left = Alignment(wrap_text=True, vertical="top", horizontal="left")


def _style(cell, value, bg=C_WHITE, bold=False, center=True, color=C_BLACK, size=10):
    cell.value = value
    cell.font = Font(name="Arial", bold=bold, size=size, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = _wrap_center if center else _wrap_top_left
    cell.border = _thin


def _get_risk_bg(level: str):
    if level == "Extremo": return C_RISK_EXTREME
    if level == "Critico": return C_RISK_CRITICAL
    if level == "Alto": return C_RISK_HIGH
    if level == "Moderado": return C_RISK_MODERATE
    if level == "Bajo": return C_RISK_LOW
    return C_RISK_MINIMAL


def _get_risk_fg(level: str):
    if level == "Extremo": return C_WHITE
    if level == "Critico": return C_WHITE
    if level == "Alto": return C_WHITE
    if level == "Moderado": return C_BLACK
    if level == "Bajo": return C_BLACK
    return C_BLACK


def export_pest_management_xlsx(
    threats: List[PestThreat],
    company_name: str,
    process_area: str,
    gate_result: dict,
    output_dir: Path
) -> PestExportResult:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MIP"

    # 1. Titulo principal
    ws.merge_cells("A1:O1")
    _style(ws["A1"], "Analisis de Riesgo del programa de Control de plagas", bg=C_TITLE_BG, bold=True, size=18, center=True)
    ws.row_dimensions[1].height = 30

    # 2. Subtitulo y fecha
    ws.merge_cells("A2:E2")
    _style(ws["A2"], f"Caracterizacion y evaluacion de riesgos - {process_area} ({company_name})", center=False, bold=True, size=11)
    ws.merge_cells("J2:O2")
    _style(ws["J2"], f"Fecha: {datetime.now().strftime('%d/%m/%Y')}", center=False, bold=True, size=11)

    # 3. Encabezados de tabla (Fila 4 y 5)
    _style(ws.cell(row=4, column=1), "Descripcion de areas operativas", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=2), "Riesgo de presencia de plagas", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=3), "Causa(s) Potencial(es) Que causa que la plaga se presente? Ej. Diseno sanitario, espacios muertos, aberturas en areas, ingreso por MP, etc.", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=4), "Enfoque estrategico", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=5), "Enfoque operativo", bg=C_HEADER_BG, bold=True, size=9, center=True)
    ws.merge_cells("F4:H4")
    _style(ws["F4"], "Evaluacion del Riesgo", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=9), "Umbral aceptable por plagas en areas", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=10), "Acciones recomendadas planta", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=11), "Acciones recomendadas proveedor", bg=C_HEADER_BG, bold=True, size=9, center=True)
    _style(ws.cell(row=4, column=12), "Responsables y fechas", bg=C_HEADER_BG, bold=True, size=9, center=True)
    ws.merge_cells("M4:O4")
    _style(ws["M4"], "Re-evaluacion del Riesgo", bg=C_HEADER_BG, bold=True, size=9, center=True)

    # Sub-encabezados (Fila 5)
    _style(ws.cell(row=5, column=6), "Sev.", bg=C_HEADER_BG, bold=True, size=9)
    _style(ws.cell(row=5, column=7), "Ocurr Bioindic", bg=C_HEADER_BG, bold=True, size=9)
    _style(ws.cell(row=5, column=8), "Nivel Riesgo", bg=C_HEADER_BG, bold=True, size=9)
    _style(ws.cell(row=5, column=13), "Sev.", bg=C_HEADER_BG, bold=True, size=9)
    _style(ws.cell(row=5, column=14), "Ocurr.", bg=C_HEADER_BG, bold=True, size=9)
    _style(ws.cell(row=5, column=15), "Nivel Riesgo", bg=C_HEADER_BG, bold=True, size=9)

    # 4. Filas de datos
    for ri, t in enumerate(threats, start=6):
        _style(ws.cell(row=ri, column=1), t.area, center=False)
        _style(ws.cell(row=ri, column=2), t.pest_risk, center=False)
        _style(ws.cell(row=ri, column=3), t.potential_causes, center=False)
        _style(ws.cell(row=ri, column=4), t.strategic_approach, center=False)
        _style(ws.cell(row=ri, column=5), t.operational_approach, center=False)

        _style(ws.cell(row=ri, column=6), t.severity)
        _style(ws.cell(row=ri, column=7), t.occurrence)
        risk_bg = _get_risk_bg(t.risk_level)
        risk_fg = _get_risk_fg(t.risk_level)
        _style(ws.cell(row=ri, column=8), t.risk_score, bg=risk_bg, bold=True, color=risk_fg)

        _style(ws.cell(row=ri, column=9), t.acceptable_threshold, center=False)
        _style(ws.cell(row=ri, column=10), t.plant_actions, center=False)
        _style(ws.cell(row=ri, column=11), t.supplier_actions, center=False)
        _style(ws.cell(row=ri, column=12), "")

        _style(ws.cell(row=ri, column=13), t.residual_severity)
        _style(ws.cell(row=ri, column=14), t.residual_occurrence)
        res_bg = _get_risk_bg(t.residual_risk_level)
        res_fg = _get_risk_fg(t.residual_risk_level)
        _style(ws.cell(row=ri, column=15), t.residual_risk_score, bg=res_bg, bold=True, color=res_fg)

        ws.row_dimensions[ri].height = 100

    # 5. Nota al pie
    last_data_row = 6 + len(threats)
    note_row = last_data_row + 1
    ws.merge_cells(f"A{note_row}:E{note_row}")
    _style(ws[f"A{note_row}"], "NR(Nivel de riesgo): Resulta de multiplicar los indices de Severidad x Ocurrencia = NR. Las operaciones que hayan obtenido los mayores valores de NR seran los primeros en aplicar Acciones recomendadas", center=False, size=9)

    # Anchos de columna (basados en MIP.xlsx)
    widths = [22, 25, 35, 30, 28, 6, 8, 10, 22, 28, 25, 14, 6, 8, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Hoja de criterios
    ws2 = wb.create_sheet("Criterios Evaluacion")
    ws2.merge_cells("A1:D1")
    _style(ws2["A1"], "Criterios de Evaluacion", bg=C_TITLE_BG, bold=True, size=14)
    ws2.row_dimensions[1].height = 30

    _style(ws2.cell(row=2, column=1), "Severidad", bg=C_TITLE_BG, bold=True, size=12)

    sev_headers = ["Grado", "Nivel", "Severidad del efecto en el proceso"]
    for ci, h in enumerate(sev_headers, 1):
        _style(ws2.cell(row=3, column=ci), h, bg=C_HEADER_BG, bold=True)

    sev_data = [
        (1, "Baja", "Contaminacion insignificante, no representa un riesgo para la salud del consumidor."),
        (2, "Moderada", "Contaminacion leve, podria causar efectos adversos menores y aislados en la salud del consumidor."),
        (3, "Alta", "Contaminacion significativa, podria causar enfermedades alimentarias moderadas en un grupo limitado de consumidores."),
        (4, "Muy Alta", "Contaminacion grave, probable causa de enfermedades alimentarias graves en un amplio grupo de consumidores."),
        (5, "Catastrofica", "Contaminacion severa, puede resultar en enfermedades alimentarias graves o incluso fatales, y afectaria a una gran cantidad de consumidores."),
    ]
    for ri, row_data in enumerate(sev_data, 4):
        for ci, val in enumerate(row_data, 1):
            _style(ws2.cell(row=ri, column=ci), val, center=False)

    # Risk matrix
    _style(ws2.cell(row=10, column=2), "Probabilidad / Severidad", bg=C_TITLE_BG, bold=True, size=11)
    prob_headers = ["", "Baja (1)", "Moderada (2)", "Alta (3)", "Muy Alta (4)", "Catastrofica (5)"]
    for ci, h in enumerate(prob_headers, 2):
        _style(ws2.cell(row=11, column=ci), h, bg=C_HEADER_BG, bold=True, center=True)

    matrix = [
        ("Muy Baja (1)", 1, 2, 3, 4, 5),
        ("Baja (2)", 2, 4, 6, 8, 10),
        ("Moderada (3)", 3, 6, 9, 12, 15),
        ("Alta (4)", 4, 8, 12, 16, 20),
        ("Muy Alta (5)", 5, 10, 15, 20, 25),
    ]

    for ri, m_row in enumerate(matrix, 12):
        for ci, val in enumerate(m_row, 2):
            cell = ws2.cell(row=ri, column=ci)
            if ci == 2:
                _style(cell, val, bg=C_HEADER_BG, bold=True, center=True)
            else:
                score = int(val)
                level = get_pest_risk_level_for_matrix(score)
                _style(cell, val, bg=_get_risk_bg(level), color=_get_risk_fg(level), bold=True)

    # Occurrence section
    ws2.merge_cells("A18:D18")
    _style(ws2.cell(row=18, column=1), "Ocurrencia", bg=C_TITLE_BG, bold=True, size=12)
    occ_headers = ["Grado", "Probabilidad de falla", "Rango de posibles fallas"]
    for ci, h in enumerate(occ_headers, 1):
        _style(ws2.cell(row=19, column=ci), h, bg=C_HEADER_BG, bold=True)

    occ_data = [
        (1, "Muy Baja", "Es poco probable que ocurra (menos del 5% de probabilidad)."),
        (2, "Baja", "Puede ocurrir en casos raros (5% - 15% de probabilidad)."),
        (3, "Moderada", "Puede ocurrir ocasionalmente (15% - 35% de probabilidad)."),
        (4, "Alta", "Es probable que ocurra (35% - 70% de probabilidad)."),
        (5, "Muy Alta", "Es muy probable que ocurra (mas del 70% de probabilidad)."),
    ]
    for ri, row_data in enumerate(occ_data, 20):
        for ci, val in enumerate(row_data, 1):
            _style(ws2.cell(row=ri, column=ci), val, center=False)

    # Sello
    stamp_row = last_data_row + 2
    ws.merge_cells(f"A{stamp_row}:O{stamp_row}")
    stamp_text = gate_result.get("xlsx_stamp", "Analisis generado por VIGIA Control de Plagas")
    _style(ws[f"A{stamp_row}"], stamp_text, color="808080", size=10, center=True)

    # Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_company = "".join(c if c.isalnum() else "_" for c in company_name)[:20]
    safe_area = "".join(c if c.isalnum() else "_" for c in process_area)[:20]
    file_name = f"control_plagas_{safe_company}_{safe_area}_{timestamp}.xlsx"
    file_path = output_dir / file_name
    wb.save(file_path)

    high_priority_rows = sum(1 for t in threats if t.risk_score >= 12)

    return PestExportResult(
        file_path=str(file_path),
        file_name=file_name,
        download_url=f"/outputs/{file_name}",
        total_rows=len(threats),
        high_priority_rows=high_priority_rows
    )


def get_pest_risk_level_for_matrix(score: int) -> str:
    if score >= 25: return "Extremo"
    if score >= 12: return "Critico"
    if score >= 6: return "Alto"
    if score >= 4: return "Moderado"
    if score >= 2: return "Bajo"
    return "Minimo"
