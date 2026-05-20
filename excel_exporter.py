"""
exporters/excel_exporter.py
Genera un archivo .xlsx con formato profesional para el plan HACCP.

Hojas:
  1. Análisis HACCP  — tabla principal, PCCs resaltados en amarillo
  2. Retiros FDA      — recalls que informaron el análisis
  3. Metadatos        — trazabilidad del reporte
"""
import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import HACCPRow, FDARecall

logger = logging.getLogger("vigia.excel_exporter")

# ── Paleta VIGÍA ──────────────────────────────────────────────────────────
C_HEADER_BG = "1F4E79"   # Azul oscuro
C_HEADER_FG = "FFFFFF"   # Blanco
C_PCC_BG    = "FFE699"   # Amarillo advertencia  (filas PCC = "Sí")
C_PCC_FG    = "C55A11"   # Naranja texto PCC
C_MON_BG    = "FFF2CC"   # Amarillo suave        (filas "monitoreo continuo")
C_STRIP_BG  = "EBF3FB"   # Azul muy claro        (filas alternas normales)
C_WHITE     = "FFFFFF"
C_BORDER    = "BDD7EE"
C_FDA_BG    = "FFF2CC"
C_META_BG   = "F2F2F2"

_thin = Border(
    left=Side(style="thin", color=C_BORDER),
    right=Side(style="thin", color=C_BORDER),
    top=Side(style="thin", color=C_BORDER),
    bottom=Side(style="thin", color=C_BORDER),
)
_wrap_left   = Alignment(wrap_text=True, vertical="top", horizontal="left")
_wrap_center = Alignment(wrap_text=True, vertical="top", horizontal="center")


def _cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col)


def _style(cell, value, bg: str, bold=False, center=False, color="000000", size=9):
    cell.value     = value
    cell.font      = Font(name="Calibri", bold=bold, size=size, color=color)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = _wrap_center if center else _wrap_left
    cell.border    = _thin


def _header(cell, text: str):
    cell.value     = text
    cell.font      = Font(name="Calibri", bold=True, size=10, color=C_HEADER_FG)
    cell.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
    cell.alignment = _wrap_center
    cell.border    = _thin


# ── Hoja 1: Análisis HACCP ────────────────────────────────────────────────

_HEADERS = [
    "Etapa del proceso",
    "Peligro identificado",
    "Medida preventiva",
    "¿PCC?",
    "Límite crítico",
    "Monitoreo",
    "Acción correctiva",
    "Verificación",
    "Registro",
]
_WIDTHS = [22, 42, 42, 28, 36, 42, 42, 42, 36]


def _row_bg(row: HACCPRow, index: int) -> str:
    """Elige el color de fondo según el tipo de PCC label."""
    if row.es_pcc:
        return C_PCC_BG
    if "monitoreo continuo" in row.pcc_label.lower():
        return C_MON_BG
    return C_STRIP_BG if index % 2 == 0 else C_WHITE


def _build_haccp_sheet(ws, product_name: str, rows: list[HACCPRow], fda_count: int):
    # Título
    ws.merge_cells("A1:I1")
    ws["A1"].value     = f"PLAN HACCP — {product_name.upper()}"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=14, color=C_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtítulo
    ws.merge_cells("A2:I2")
    ws["A2"].value = (
        f"Generado por VIGÍA HACCP MCP  |  "
        f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
        f"Retiros FDA consultados: {fda_count}  |  "
        f"Normativa: NOM-251-SSA1-2009 / Codex CAC/RCP 1-1969"
    )
    ws["A2"].font      = Font(name="Calibri", italic=True, size=9, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6   # separador

    # Encabezados
    for ci, h in enumerate(_HEADERS, 1):
        _header(ws.cell(row=4, column=ci), h)
    ws.row_dimensions[4].height = 34

    # Filas de datos
    for ri, row in enumerate(rows, start=5):
        bg = _row_bg(row, ri)

        # Color de texto para la celda ¿PCC?
        pcc_color = C_PCC_FG if row.es_pcc else "404040"
        pcc_bold  = row.es_pcc

        values = [
            row.etapa,
            row.peligro,
            row.medida_preventiva,
            row.pcc_label,          # ← etiqueta textual matizada
            row.limite_critico,
            row.monitoreo,
            row.accion_correctiva,
            row.verificacion,
            row.registro,
        ]
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=ri, column=ci)
            if ci == 4:   # columna ¿PCC?
                _style(c, val, bg, bold=pcc_bold, center=True, color=pcc_color)
            else:
                _style(c, val, bg)

        ws.row_dimensions[ri].height = 78

    # Anchos
    for ci, w in enumerate(_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Resumen PCCs debajo de la tabla
    last_row = 4 + len(rows) + 2
    pcc_list = [r.etapa for r in rows if r.es_pcc]
    ws.merge_cells(f"A{last_row}:I{last_row}")
    ws[f"A{last_row}"].value = (
        f"PCCs identificados ({len(pcc_list)}): "
        + ("  |  ".join(pcc_list) if pcc_list else "Ninguno en este flujo de proceso")
    )
    ws[f"A{last_row}"].font      = Font(name="Calibri", bold=True, size=9, color=C_HEADER_BG)
    ws[f"A{last_row}"].alignment = _wrap_left
    ws.row_dimensions[last_row].height = 20

    ws.freeze_panes = "A5"


# ── Hoja 2: Retiros FDA ───────────────────────────────────────────────────

_FDA_HEADERS = ["N° Retiro", "Empresa", "Producto", "Motivo", "Fecha", "Clasificación", "Estado"]
_FDA_WIDTHS  = [14, 28, 44, 54, 13, 13, 12]


def _build_fda_sheet(ws, recalls: list[FDARecall]):
    ws.merge_cells("A1:G1")
    ws["A1"].value     = "RETIROS DE MERCADO FDA — Datos que informaron el análisis HACCP"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=12, color=C_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    for ci, h in enumerate(_FDA_HEADERS, 1):
        _header(ws.cell(row=2, column=ci), h)
    ws.row_dimensions[2].height = 24

    if not recalls:
        ws.merge_cells("A3:G3")
        ws["A3"].value = "No se encontraron retiros de mercado relevantes en el período consultado."
        ws["A3"].font  = Font(italic=True, color="595959")
        return

    for ri, r in enumerate(recalls, start=3):
        bg = C_FDA_BG if ri % 2 == 0 else C_WHITE
        for ci, val in enumerate([
            r.recall_number,
            r.recalling_firm,
            r.product_description[:140],
            r.reason_for_recall[:200],
            r.recall_initiation_date,
            r.classification,
            r.status,
        ], 1):
            _style(ws.cell(row=ri, column=ci), val, bg, size=9)
        ws.row_dimensions[ri].height = 50

    for ci, w in enumerate(_FDA_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ── Hoja 3: Metadatos ─────────────────────────────────────────────────────

def _build_metadata_sheet(ws, product_name: str, fda_hazards: set[str], step_count: int):
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 72

    rows = [
        ("Plataforma",              "VIGÍA HACCP MCP v1.0"),
        ("Producto analizado",      product_name),
        ("Etapas del proceso",      str(step_count)),
        ("Peligros FDA detectados", ", ".join(sorted(fda_hazards)) or "Ninguno"),
        ("Normativas de referencia","NOM-251-SSA1-2009 · NOM-213-SSA1-2018 · Codex CAC/RCP 1-1969 Rev.4"),
        ("Estándar HACCP",          "Codex Alimentarius (7 principios, 12 pasos)"),
        ("Fuente de datos FDA",     "openFDA Food Enforcement API (últimos 5 años)"),
        ("Fecha de generación",     datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("Advertencia",
         "Este análisis es una herramienta de soporte técnico. "
         "Debe ser revisado y validado por un consultor HACCP certificado "
         "antes de su implementación en planta."),
    ]
    for i, (k, v) in enumerate(rows, 1):
        ck = ws.cell(row=i, column=1)
        cv = ws.cell(row=i, column=2)
        ck.value = k
        cv.value = v
        ck.font  = Font(name="Calibri", bold=True, size=10)
        cv.font  = Font(name="Calibri", size=10)
        ck.fill  = PatternFill("solid", fgColor=C_META_BG)
        ck.alignment = _wrap_left
        cv.alignment = _wrap_left
        ck.border    = _thin
        cv.border    = _thin
        ws.row_dimensions[i].height = 22


# ── Punto de entrada ──────────────────────────────────────────────────────

def export_to_excel(
    product_name: str,
    haccp_rows:   list[HACCPRow],
    fda_recalls:  list[FDARecall],
    fda_hazards:  set[str],
    output_dir:   Path,
) -> Path:
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Análisis HACCP"
    _build_haccp_sheet(ws1, product_name, haccp_rows, len(fda_recalls))

    ws2 = wb.create_sheet("Retiros FDA")
    _build_fda_sheet(ws2, fda_recalls)

    ws3 = wb.create_sheet("Metadatos")
    _build_metadata_sheet(ws3, product_name, fda_hazards, len(haccp_rows))

    wb.active = ws1

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe      = "".join(c if c.isalnum() else "_" for c in product_name)[:30]
    path      = output_dir / f"HACCP_{safe}_{timestamp}.xlsx"
    wb.save(path)
    logger.info("Excel guardado: %s", path)
    return path
